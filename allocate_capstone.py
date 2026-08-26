#!/usr/bin/env python3
"""Allocate capstone students to client projects from the interest survey."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import pulp
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKBOOK = Path(__file__).resolve().parent / "Capstone Project Interest Survey (Fall 2026).xlsx"
SHEET_IN = "Sheet1"
SHEET_V1 = "Allocations"
SHEET_V2 = "Allocations - PS 50-50"
SHEET_V3 = "Allocations - tech mix"
SHEET_V4 = "Allocations - gender agnostic"
SHEET_V5 = "Allocations - gender mix"
SHEET_V6 = "Allocations - no lone female"
SHEET_V7 = "Agnostic + Amitha"
SHEET_V8 = "Gender mix + Amitha"
SHEET_V9 = "No lone female + Amitha"
SHEET_V10 = "Allocations - even gender"
SHEET_V11 = "Even gender + Amitha"
SHEET_V12 = "Allocations - even + mix"
SHEET_V13 = "Even + mix + Amitha"
ROSTER_MSAI = Path(__file__).resolve().parent / "roster_msai.xlsx"
ROSTER_MBAI = Path(__file__).resolve().parent / "roster_mbai.xlsx"

UNLISTED_COST = 100
SIZE4_TEAMS = 4
SIZE5_TEAMS = 13
SIZE6_TEAMS = 0

AVOID_PAIRS = [
    ("Jordan Johnson", "Jianchen Hong"),
    ("Jordan Johnson", "Allan Tang"),
    ("Samyak Jain", "Deepesh Khubchandani"),
    ("Samyak Jain", "Tanu Sharma"),
    ("Karl Morcott", "Daniela Paredes Hosage"),
]
PROJECT_BANS = {
    "Simron Patel": ["Grainger"],
}
FORCED = {
    "Erik Beitel": "Baxter",
}
FORCED_WHITE = {"Erik Beitel"}
UNROSTERED_GENDER = {
    "Amitha Javare Gowda": "f",  # not on class roster; inferred from given name
}
MERGE_BY_NAME = {"vandan agrawal"}

# 3 = more tech (lean MSAI on 5-person teams); 1 = more business (lean MBAi).
TECH_RATING = {
    "Abbott": 2,
    "Accenture": 3,
    "Baxter": 2,
    "CDW": 2,
    "Deere": 3,
    "Deloitte": 2,
    "Disney": 2,
    "EY": 2,
    "Feinberg Oncology": 1,
    "Future Energy Enterprises": 2,
    "Grainger": 2,
    "Home Depot": 3,
    "Per Scholas": 2,
    "Point Taken": 1,
    "Prologis": 2,
    "Vanguard": 2,
    "ViewSonic": 2,
}

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFE699")
ORANGE = PatternFill("solid", fgColor="F4B183")
GRAY = PatternFill("solid", fgColor="D9D9D9")
RED = PatternFill("solid", fgColor="FF7C80")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
WILDCARD_FILL = PatternFill("solid", fgColor="A3E4D7")
MALE_FILL = PatternFill("solid", fgColor="5D6D7E")
FEMALE_FILL = PatternFill("solid", fgColor="AF7AC5")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
COMPANY_FILL = PatternFill("solid", fgColor="D6EAF8")
SPACER_FILL = PatternFill("solid", fgColor="E8EEF2")
MSAI_FILL = PatternFill("solid", fgColor="1B6CA8")
MBAI_FILL = PatternFill("solid", fgColor="8E4585")
WHITE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BOLD = Font(bold=True, name="Calibri", size=11)
NORMAL = Font(name="Calibri", size=11)
FOOTER_FONT = Font(name="Calibri", size=11, color="333333")
MSAI_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
MBAI_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def parse_points(value) -> int:
    if value is None:
        return 0
    match = re.search(r"-?\d+", str(value).strip())
    return int(match.group()) if match else 0


def short_program(program: str) -> str:
    if "MSAI" in program:
        return "MSAI"
    if "MBAi" in program or "MBAI" in program.upper():
        return "MBAi"
    return program


def project_name(header: str) -> str:
    return str(header).split("\xa0.")[-1].strip()


def load_students(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET_IN]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    proj_cols = []
    for idx, header in enumerate(headers):
        if header and "allocating a total of 20 POINTS" in str(header):
            proj_cols.append((idx, project_name(header)))

    active = [name for _, name in proj_cols if "cancel" not in name.lower()]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or not raw[5]:
            continue
        first = str(raw[5]).strip()
        last = str(raw[6]).strip() if raw[6] else ""
        bids = {name: parse_points(raw[idx]) for idx, name in proj_cols}
        cancelled = {name: pts for name, pts in bids.items() if "cancel" in name.lower() and pts}
        active_bids = {name: bids.get(name, 0) for name in active}
        rows.append(
            {
                "id": raw[0],
                "complete": raw[2],
                "first": first,
                "last": last,
                "name": f"{first} {last}".strip(),
                "preferred": (str(raw[7]).strip() if raw[7] else ""),
                "email": (str(raw[8]).strip() if raw[8] else ""),
                "program": short_program(str(raw[10]).strip() if raw[10] else ""),
                "active_bids": active_bids,
                "cancelled_points": sum(cancelled.values()),
                "form_total": sum(bids.values()),
            }
        )

    grouped = defaultdict(list)
    for row in rows:
        key = row["name"].strip().lower() if row["name"].strip().lower() in MERGE_BY_NAME else row["email"].lower()
        grouped[key].append(row)

    students = []
    for key, group in grouped.items():
        group.sort(key=lambda r: (r["complete"] is not None, r["complete"] or 0))
        students.append(group[-1])
    students.sort(key=lambda s: (s["last"].lower(), s["first"].lower(), s["email"].lower()))
    return students, active


def load_roster(path: Path, program: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    people = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or not raw[0]:
            continue
        email = str(raw[0]).strip()
        roster_name = str(raw[1]).strip()
        gender = str(raw[2]).strip().lower()
        last, first = (roster_name.split(",", 1) + [""])[:2]
        people.append(
            {
                "program": program,
                "email": email,
                "roster_name": roster_name,
                "last": last.strip(),
                "first": first.strip(),
                "gender": gender,
            }
        )
    return people


def _email_keys(email: str) -> set[str]:
    e = email.strip().lower().replace("kelllogg", "kellogg")
    local, _, domain = e.partition("@")
    keys = {e, f"{local.replace('.', '')}@{domain}"}
    return keys


def _name_key(last: str, first: str) -> str:
    first_token = first.strip().split()[0].lower() if first.strip() else ""
    return f"{last.strip().lower()}|{first_token}"


def merge_roster_and_survey(survey_students, projects, include_unrostered=False):
    """Roster is the population. Survey bids attach by email/name. Unmatched roster = wildcards."""
    roster = load_roster(ROSTER_MSAI, "MSAI") + load_roster(ROSTER_MBAI, "MBAi")
    survey_by_key = {}
    for s in survey_students:
        for key in _email_keys(s["email"]):
            survey_by_key.setdefault(key, s)
    survey_by_name = {}
    for s in survey_students:
        survey_by_name.setdefault(_name_key(s["last"], s["first"]), []).append(s)

    used_survey_ids = set()
    merged = []
    unmatched_roster = []
    for r in roster:
        hit = None
        for key in _email_keys(r["email"]):
            if key in survey_by_key:
                hit = survey_by_key[key]
                break
        if hit is None:
            name_hits = survey_by_name.get(_name_key(r["last"], r["first"]), [])
            name_hits = [s for s in name_hits if id(s) not in used_survey_ids]
            if len(name_hits) == 1:
                hit = name_hits[0]
        if hit is None:
            unmatched_roster.append(r)
            display_first = r["first"].split()[0] if r["first"] else ""
            merged.append(
                {
                    "id": None,
                    "complete": None,
                    "first": r["first"],
                    "last": r["last"],
                    "name": f"{display_first} {r['last']}".strip(),
                    "preferred": "",
                    "email": r["email"],
                    "program": r["program"],
                    "gender": r["gender"],
                    "active_bids": {p: 0 for p in projects},
                    "cancelled_points": 0,
                    "form_total": 0,
                    "wildcard": True,
                }
            )
            continue
        used_survey_ids.add(id(hit))
        rec = dict(hit)
        rec["gender"] = r["gender"]
        rec["wildcard"] = False
        rec["email"] = r["email"]  # canonical roster email
        merged.append(rec)

    leftover = [s for s in survey_students if id(s) not in used_survey_ids]
    if include_unrostered:
        for s in leftover:
            rec = dict(s)
            rec["gender"] = UNROSTERED_GENDER.get(s["name"])
            if rec["gender"] not in ("m", "f"):
                raise SystemExit(f"Need a gender for unrostered student {s['name']!r}")
            rec["wildcard"] = False
            rec["unrostered"] = True
            merged.append(rec)
        leftover = []
    merged.sort(key=lambda s: (s["last"].lower(), s["first"].lower(), s["email"].lower()))
    return merged, unmatched_roster, leftover


def choose_team_sizes(n_students: int, n_projects: int) -> tuple[int, int, int]:
    extra = n_students - 4 * n_projects
    if extra < 0:
        raise SystemExit(f"Not enough students ({n_students}) for {n_projects} teams of 4+")
    for sixes in range(0, extra // 2 + 1):
        fives = extra - 2 * sixes
        fours = n_projects - fives - sixes
        if fours >= 0 and fives >= 0:
            return fours, fives, sixes
    raise SystemExit(f"No 4/5/6 mix for {n_students} students and {n_projects} projects")


def ranks_for(bids: dict[str, int]) -> dict[str, int | None]:
    """Dense rank by distinct positive point values. None = unlisted."""
    positive = {p: v for p, v in bids.items() if v > 0}
    values = sorted(set(positive.values()), reverse=True)
    value_to_rank = {val: i + 1 for i, val in enumerate(values)}
    return {p: (value_to_rank[v] if v > 0 else None) for p, v in bids.items()}


def penalty(rank: int | None) -> int:
    if rank is None:
        return UNLISTED_COST
    return (rank - 1) ** 2


def pref_label(rank: int | None, wildcard: bool = False) -> str:
    if wildcard:
        return "Wildcard"
    if rank is None:
        return "Not listed"
    if rank == 1:
        return "1st"
    if rank == 2:
        return "2nd"
    if rank == 3:
        return "3rd"
    return f"{rank}th"


def pref_fill(rank: int | None, wildcard: bool = False) -> PatternFill:
    if wildcard:
        return WILDCARD_FILL
    if rank is None:
        return RED
    if rank == 1:
        return GREEN
    if rank == 2:
        return YELLOW
    if rank == 3:
        return ORANGE
    return GRAY


def find_student(students, full_name: str):
    matches = [s for s in students if s["name"] == full_name]
    if len(matches) != 1:
        raise SystemExit(f"Expected 1 student named {full_name!r}, found {len(matches)}")
    return matches[0]


def allocate(students, projects, balanced_projects=(), use_tech_mix=False, forbid_2m1f=False, forbid_lone_female=False, even_gender=False, forced=None):
    if forced is None:
        forced = FORCED

    names = [s["name"] for s in students]
    if len(names) != len(set(names)):
        dupes = [n for n in names if names.count(n) > 1]
        raise SystemExit(f"Duplicate names after dedup: {sorted(set(dupes))}")

    msai = [s["name"] for s in students if s["program"] == "MSAI"]
    mbai = [s["name"] for s in students if s["program"] == "MBAi"]
    if len(msai) + len(mbai) != len(students):
        raise SystemExit(f"Unexpected programs: {sorted({s['program'] for s in students})}")

    ranks = {s["name"]: ranks_for(s["active_bids"]) for s in students}
    cost = {}
    for s in students:
        for p in projects:
            if s.get("wildcard"):
                cost[s["name"], p] = 0
            else:
                cost[s["name"], p] = penalty(ranks[s["name"]][p])

    n = len(students)
    n4, n5, n6 = choose_team_sizes(n, len(projects))

    prob = pulp.LpProblem("capstone_allocation", pulp.LpMinimize)
    x = pulp.LpVariable.dicts("assign", (names, projects), cat="Binary")
    is4 = pulp.LpVariable.dicts("size4", projects, cat="Binary")
    is5 = pulp.LpVariable.dicts("size5", projects, cat="Binary")
    is6 = pulp.LpVariable.dicts("size6", projects, cat="Binary")
    unbal4 = pulp.LpVariable.dicts("unbal4", projects, cat="Binary")

    prob += (
        pulp.lpSum(cost[i, j] * x[i][j] for i in names for j in projects)
        + 8 * pulp.lpSum(unbal4[j] for j in projects)
    )

    for i in names:
        prob += pulp.lpSum(x[i][j] for j in projects) == 1, f"one_project_{i}"

    for j in projects:
        prob += is4[j] + is5[j] + is6[j] == 1, f"one_size_{j}"
        size = 4 * is4[j] + 5 * is5[j] + 6 * is6[j]
        prob += pulp.lpSum(x[i][j] for i in names) == size, f"team_size_{j}"
        msai_count = pulp.lpSum(x[i][j] for i in msai)
        five_lo, five_hi = 2, 3
        if use_tech_mix:
            rating = TECH_RATING[j]
            if rating == 3:
                five_lo, five_hi = 3, 3
            elif rating == 1:
                five_lo, five_hi = 2, 2
        prob += msai_count >= 1 * is4[j] + five_lo * is5[j] + 3 * is6[j], f"msai_lo_{j}"
        prob += msai_count <= 3 * is4[j] + five_hi * is5[j] + 3 * is6[j], f"msai_hi_{j}"
        prob += msai_count <= 2 + unbal4[j] + 3 * (1 - is4[j]), f"unbal4_hi_{j}"
        prob += msai_count >= 2 - unbal4[j] - 3 * (1 - is4[j]), f"unbal4_lo_{j}"

    prob += pulp.lpSum(is4[j] for j in projects) == n4
    prob += pulp.lpSum(is5[j] for j in projects) == n5
    prob += pulp.lpSum(is6[j] for j in projects) == n6

    for person, company in forced.items():
        find_student(students, person)
        if company not in projects:
            raise SystemExit(f"Forced project missing: {company}")
        prob += x[person][company] == 1, f"force_{person}"

    for person, banned in PROJECT_BANS.items():
        find_student(students, person)
        for company in banned:
            prob += x[person][company] == 0, f"ban_{person}_{company}"

    for a, b in AVOID_PAIRS:
        find_student(students, a)
        find_student(students, b)
        for j in projects:
            prob += x[a][j] + x[b][j] <= 1, f"avoid_{a}_{b}_{j}"

    for project in balanced_projects:
        if project not in projects:
            raise SystemExit(f"Balanced project missing: {project}")
        prob += is5[project] == 0, f"even_size_{project}"
        prob += pulp.lpSum(x[i][project] for i in msai) == pulp.lpSum(x[i][project] for i in mbai), f"fifty_{project}"
        prob += unbal4[project] == 0, f"bal_unbal_{project}"

    if forbid_2m1f:
        # A 3-person cohort slice may be 3M, 3F, or 1M+2F — not 2M+1F.
        for cohort_name, cohort in (("MSAI", msai), ("MBAi", mbai)):
            females = [s["name"] for s in students if s["name"] in cohort and s.get("gender") == "f"]
            males = [s["name"] for s in students if s["name"] in cohort and s.get("gender") == "m"]
            if len(females) + len(males) != len(cohort):
                raise SystemExit(f"Missing gender in {cohort_name}")
            for j in projects:
                n_c = pulp.lpSum(x[i][j] for i in cohort)
                n_f = pulp.lpSum(x[i][j] for i in females)
                n_m = pulp.lpSum(x[i][j] for i in males)
                is3 = pulp.LpVariable(f"is3_{cohort_name}_{j}", cat="Binary")
                all_m3 = pulp.LpVariable(f"allm3_{cohort_name}_{j}", cat="Binary")
                prob += n_c >= 3 * is3
                prob += n_c <= 2 + is3
                prob += all_m3 <= is3
                # if exactly 3 and not all-male: at most 1 male (so 1M2F or 3F)
                prob += n_m <= 1 + 2 * (1 - is3) + 2 * all_m3
                prob += n_m >= 3 * all_m3
                prob += n_f <= 3 * (1 - all_m3)

    if forbid_lone_female:
        females = [s["name"] for s in students if s.get("gender") == "f"]
        for j in projects:
            n_f = pulp.lpSum(x[i][j] for i in females)
            has_fs = pulp.LpVariable(f"has2f_{j}", cat="Binary")
            # 0 females, or at least 2 — never exactly 1
            prob += n_f >= 2 * has_fs
            prob += n_f <= 6 * has_fs

    if even_gender:
        # Spread women as evenly as possible without a lone female on any team.
        # Odd woman-count: exactly one 3F team, the rest 0F or 2F.
        # Even woman-count: every team is 0F or 2F.
        females = [s["name"] for s in students if s.get("gender") == "f"]
        n_triple = 0 if len(females) % 2 == 0 else 1
        is3f = pulp.LpVariable.dicts("even_is3f", projects, cat="Binary")
        for j in projects:
            n_f = pulp.lpSum(x[i][j] for i in females)
            has_fs = pulp.LpVariable(f"even_has2f_{j}", cat="Binary")
            prob += n_f >= 2 * has_fs, f"even_f_lo_{j}"
            prob += n_f <= 2 * has_fs + is3f[j], f"even_f_hi_{j}"
            prob += is3f[j] <= has_fs, f"even_3f_needs_women_{j}"
            prob += n_f <= 2 + 3 * (1 - is4[j]), f"even_f_no3_on4_{j}"
            prob += is3f[j] <= 1 - is4[j], f"even_no3_on4team_{j}"
        prob += pulp.lpSum(is3f[j] for j in projects) == n_triple, "even_n_triple"

    solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=120, gapRel=0)
    status = prob.solve(solver)
    if pulp.LpStatus[status] != "Optimal":
        raise SystemExit(f"Solver status: {pulp.LpStatus[status]}")

    assignment = {}
    teams = {p: [] for p in projects}
    for s in students:
        chosen = [p for p in projects if pulp.value(x[s["name"]][p]) > 0.5]
        if len(chosen) != 1:
            raise SystemExit(f"Bad assignment for {s['name']}: {chosen}")
        project = chosen[0]
        rank = ranks[s["name"]][project]
        rec = {
            **s,
            "project": project,
            "rank": rank,
            "points": s["active_bids"][project],
            "cost": 0 if s.get("wildcard") else penalty(rank),
        }
        assignment[s["name"]] = rec
        teams[project].append(rec)

    for person, company in forced.items():
        if assignment[person]["project"] != company:
            raise SystemExit(f"Force failed: {person}")
    for person, banned in PROJECT_BANS.items():
        if assignment[person]["project"] in banned:
            raise SystemExit(f"Ban failed: {person}")
    for a, b in AVOID_PAIRS:
        if assignment[a]["project"] == assignment[b]["project"]:
            raise SystemExit(f"Avoid failed: {a} / {b}")
    for project, members in teams.items():
        size = len(members)
        n_msai = sum(1 for m in members if m["program"] == "MSAI")
        n_mbai = size - n_msai
        if size == 4 and (min(n_msai, n_mbai) < 1 or max(n_msai, n_mbai) > 3):
            raise SystemExit(f"{project} 4-person mix {n_msai}/{n_mbai}")
        if size == 5:
            if max(n_msai, n_mbai) > 3:
                raise SystemExit(f"{project} 5-person mix {n_msai}/{n_mbai}")
            if use_tech_mix:
                rating = TECH_RATING[project]
                if rating == 3 and n_msai != 3:
                    raise SystemExit(f"{project} tech-3 5-person mix {n_msai}/{n_mbai}")
                if rating == 1 and n_msai != 2:
                    raise SystemExit(f"{project} tech-1 5-person mix {n_msai}/{n_mbai}")
        if size == 6 and (n_msai, n_mbai) != (3, 3):
            raise SystemExit(f"{project} 6-person mix {n_msai}/{n_mbai}")
        if size not in (4, 5, 6):
            raise SystemExit(f"{project} has size {size}")
        if project in balanced_projects and n_msai != n_mbai:
            raise SystemExit(f"{project} not 50-50: {n_msai}/{n_mbai}")
        if forbid_2m1f:
            for cohort in ("MSAI", "MBAi"):
                slice_ = [m for m in members if m["program"] == cohort]
                if len(slice_) != 3:
                    continue
                n_f = sum(1 for m in slice_ if m.get("gender") == "f")
                n_m = len(slice_) - n_f
                if (n_m, n_f) == (2, 1):
                    raise SystemExit(f"{project} {cohort} is 2M/1F")
        if forbid_lone_female:
            n_f = sum(1 for m in members if m.get("gender") == "f")
            if n_f == 1:
                raise SystemExit(f"{project} has a lone female")
        if even_gender:
            n_f = sum(1 for m in members if m.get("gender") == "f")
            n_women = sum(1 for s in students if s.get("gender") == "f")
            max_f = 2 if n_women % 2 == 0 else 3
            if n_f == 1 or n_f > max_f:
                raise SystemExit(f"{project} female count {n_f} not in {{0, 2" + (f", {max_f}" if max_f > 2 else "") + "}")
            if size == 4 and n_f == 3:
                raise SystemExit(f"{project} 4-person team should not be 3F/1M")
    if even_gender:
        n_women = sum(1 for s in students if s.get("gender") == "f")
        n_triple = sum(
            1 for members in teams.values()
            if sum(1 for m in members if m.get("gender") == "f") >= 3
        )
        expected_triple = 0 if n_women % 2 == 0 else 1
        if n_triple != expected_triple:
            raise SystemExit(f"Even gender expected {expected_triple} triple-F team(s), got {n_triple}")

    pref_objective = sum(assignment[s["name"]]["cost"] for s in students)
    return assignment, teams, pref_objective

def bid_note(student) -> str:
    if student.get("wildcard"):
        return "No bid (wildcard; not on survey)"
    active_total = sum(student["active_bids"].values())
    parts = []
    if active_total != 20:
        parts.append(f"Active bids sum to {active_total}, not 20")
    if student["cancelled_points"]:
        parts.append(f"{student['cancelled_points']} points on cancelled Abbvie")
    if student["form_total"] != 20 and student["form_total"] != active_total:
        parts.append(f"form total {student['form_total']}")
    return "; ".join(parts)


def program_fill(program: str) -> PatternFill:
    return MSAI_FILL if program == "MSAI" else MBAI_FILL


def program_font(program: str) -> Font:
    return MSAI_FONT if program == "MSAI" else MBAI_FONT


def pref_stats(teams):
    rank_counts = defaultdict(int)
    n_wildcard = 0
    for members in teams.values():
        for m in members:
            if m.get("wildcard"):
                n_wildcard += 1
            else:
                rank_counts[m["rank"]] += 1
    return (
        rank_counts[1],
        rank_counts[2],
        rank_counts[3],
        sum(c for r, c in rank_counts.items() if r is not None and r >= 4),
        rank_counts[None],
        n_wildcard,
    )


def size_mix(teams):
    fours = [p for p, m in teams.items() if len(m) == 4]
    fives = [p for p, m in teams.items() if len(m) == 5]
    sixes = [p for p, m in teams.items() if len(m) == 6]
    return fours, fives, sixes


def teams_from_assignment_map(students, name_to_project, projects):
    ranks = {s["name"]: ranks_for(s["active_bids"]) for s in students}
    teams = {p: [] for p in projects}
    for s in students:
        project = name_to_project[s["name"]]
        rank = ranks[s["name"]][project]
        rec = {
            **s,
            "project": project,
            "rank": rank,
            "points": s["active_bids"][project],
            "cost": 0 if s.get("wildcard") else penalty(rank),
        }
        teams[project].append(rec)
    return teams


def read_assignment_map(path: Path, sheet_name: str) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    mapping = {}
    company = None
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0]:
            raw = str(row[0]).strip()
            if " (tech " in raw:
                company = raw.split(" (tech ")[0].strip()
            else:
                company = raw.split("  (")[0].split(" (")[0].strip()
        name = row[1]
        if name and company:
            mapping[str(name).strip()] = company
    return mapping


def write_sheet(
    path: Path,
    sheet_name: str,
    title: str,
    students,
    projects,
    teams,
    objective,
    footer_paragraphs,
    show_tech=False,
    show_gender=False,
):
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    n1, n2, n3, n4p, n_unlisted, n_wildcard = pref_stats(teams)
    fours, fives, sixes = size_mix(teams)
    last_col = 7 if show_gender else 6
    end_letter = get_column_letter(last_col)

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, name="Calibri", size=16, color="1F4E79")
    ws.merge_cells(f"A1:{end_letter}1")

    summary = (
        f"{len(students)} students · {len(projects)} companies · "
        f"{len(fours)} teams of 4, {len(fives)} teams of 5"
        + (f", {len(sixes)} teams of 6" if sixes else "")
        + f" · objective {objective} · "
        f"{n1} first / {n2} second / {n3} third / {n4p} 4th+ listed / {n_unlisted} unlisted"
        + (f" / {n_wildcard} wildcard" if n_wildcard else "")
    )
    ws["A2"] = summary
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="555555")
    ws.merge_cells(f"A2:{end_letter}2")

    legend = (
        "Preference: green = 1st · yellow = 2nd · orange = 3rd · gray = 4th+ listed · red = not listed · teal = wildcard · white = forced (Erik Beitel).  "
        "Program: blue = MSAI · purple = MBAi."
    )
    if show_gender:
        legend += "  Gender: slate = M · lilac = F."
    ws["A3"] = legend
    ws["A3"].font = Font(name="Calibri", size=10, color="555555")
    ws.merge_cells(f"A3:{end_letter}3")

    headers = ["Company", "Member", "Program"]
    if show_gender:
        headers.append("Gender")
    headers += ["Preference", "Points bid on this company", "Notes"]
    header_row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    pref_col = 5 if show_gender else 4
    pts_col = pref_col + 1
    notes_col = pts_col + 1

    row = header_row + 1
    for idx, project in enumerate(projects):
        members = sorted(
            teams[project],
            key=lambda m: (
                0 if m["program"] == "MSAI" else 1,
                0 if m.get("gender") == "f" else 1,
                1000 if m.get("wildcard") else (999 if m["rank"] is None else m["rank"]),
                m["last"].lower(),
                m["first"].lower(),
            ),
        )
        start = row
        for member in members:
            rank = member["rank"]
            is_wild = bool(member.get("wildcard"))
            is_forced_white = member["name"] in FORCED_WHITE
            name_fill = WHITE_FILL if is_forced_white else pref_fill(rank, is_wild)
            ws.cell(row, 2, member["name"]).fill = name_fill
            prog_cell = ws.cell(row, 3, member["program"])
            prog_cell.fill = program_fill(member["program"])
            prog_cell.font = program_font(member["program"])
            if show_gender:
                g = member.get("gender")
                g_cell = ws.cell(row, 4, "F" if g == "f" else "M" if g == "m" else "?")
                g_cell.fill = FEMALE_FILL if g == "f" else MALE_FILL
                g_cell.font = WHITE_FONT
            pref_cell = ws.cell(row, pref_col, pref_label(rank, is_wild))
            pref_cell.fill = WHITE_FILL if is_forced_white else pref_fill(rank, is_wild)
            ws.cell(row, pts_col, member["points"] if member["points"] else "—")
            ws.cell(row, notes_col, bid_note(member))
            center_cols = {3, pref_col, pts_col}
            if show_gender:
                center_cols.add(4)
            for col in range(1, last_col + 1):
                cell = ws.cell(row, col)
                if col not in (3, 4) or (col == 4 and not show_gender):
                    cell.font = BOLD if col == 2 else NORMAL
                cell.border = THIN
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                    horizontal="center" if col in center_cols else "left",
                )
            row += 1
        end = row - 1
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        rating = TECH_RATING[project]
        company_cell = ws.cell(start, 1, f"{project} (tech {rating})")
        company_cell.font = BOLD
        company_cell.fill = COMPANY_FILL
        company_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in range(start, end + 1):
            ws.cell(r, 1).border = THIN
            ws.cell(r, 1).fill = COMPANY_FILL

        if idx < len(projects) - 1:
            for col in range(1, last_col + 1):
                cell = ws.cell(row, col)
                cell.fill = SPACER_FILL
                cell.border = Border()
                cell.value = None
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
            ws.row_dimensions[row].height = 8
            row += 1

    last_data_row = row - 1
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:{end_letter}{last_data_row}"
    widths = {1: 28, 2: 28, 3: 12}
    if show_gender:
        widths.update({4: 10, 5: 16, 6: 28, 7: 48})
    else:
        widths.update({4: 16, 5: 28, 6: 48})
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[header_row].height = 28
    ws.sheet_view.showGridLines = False

    footer_row = last_data_row + 2
    for paragraph in footer_paragraphs:
        cell = ws.cell(footer_row, 1, paragraph)
        cell.font = FOOTER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=last_col)
        ws.row_dimensions[footer_row].height = 48 if len(paragraph) < 180 else 66
        footer_row += 1

    wb.save(path)

def tech_mix_note(teams) -> str:
    tech3, tech1, med_msai_heavy, med_mbai_heavy = [], [], [], []
    for project, members in teams.items():
        if len(members) != 5:
            continue
        rating = TECH_RATING[project]
        n_msai = sum(1 for m in members if m["program"] == "MSAI")
        label = f"{project} {n_msai}/{len(members)-n_msai}"
        if rating == 3:
            tech3.append(label)
        elif rating == 1:
            tech1.append(label)
        elif n_msai == 3:
            med_msai_heavy.append(label)
        else:
            med_mbai_heavy.append(label)
    return (
        "5-person mix rule: tech 3 → 3 MSAI / 2 MBAi; tech 1 → 2 MSAI / 3 MBAi; tech 2 → either. "
        f"Tech 3: {'; '.join(tech3) or 'none'}. "
        f"Tech 1: {'; '.join(tech1) or 'none'}. "
        f"Tech 2 with 3 MSAI: {'; '.join(med_msai_heavy) or 'none'}. "
        f"Tech 2 with 3 MBAi: {'; '.join(med_mbai_heavy) or 'none'}."
    )


def describe_run(teams, extra="", forced=None):
    n1, n2, n3, n4p, n_unlisted, n_wildcard = pref_stats(teams)
    fours, fives, _ = size_mix(teams)
    gray, red, orange, wild = [], [], [], []
    for project, members in teams.items():
        for m in members:
            loc = f"{m['name']} ({project})"
            if m.get("wildcard"):
                wild.append(loc)
            elif m["name"] in FORCED_WHITE:
                continue
            elif m["rank"] is None:
                red.append(loc)
            elif m["rank"] >= 4:
                gray.append(loc)
            elif m["rank"] == 3:
                orange.append(loc)
    fours_txt = ", ".join(fours) if fours else "none"
    forced = FORCED if forced is None else forced
    forced_txt = "; ".join(
        f"{person} is forced to {company} and shown in white (not red)"
        for person, company in forced.items()
    )
    lines = [
        f"Fit: {sum(len(m) for m in teams.values())} students, all {len(teams)} companies. "
        f"{len(fives)} teams of 5, {len(fours)} teams of 4 ({fours_txt}).",
        f"Preferences: {n1} first, {n2} second, {n3} third, {n4p} fourth-or-worse listed, {n_unlisted} unlisted"
        + (f", {n_wildcard} wildcard" if n_wildcard else "")
        + ".",
        "Red (not listed): " + ("; ".join(red) if red else "none") + ". "
        "Gray (4th+ listed): " + ("; ".join(gray) if gray else "none") + ". "
        "Orange (3rd): " + ("; ".join(orange) if orange else "none") + "."
        + ((" Wildcard: " + "; ".join(wild) + ".") if wild else ""),
        "Avoids are honored, including Simron off Grainger. "
        + (forced_txt + ". " if forced_txt else "")
        + "Vandan is the later submission.",
    ]
    if extra:
        lines.append(extra)
    return lines

def print_run(label, teams, objective):
    counts = defaultdict(int)
    for members in teams.values():
        for rec in members:
            counts[rec["rank"]] += 1
    print(f"\n=== {label} ===")
    print(f"Objective: {objective}")
    print("Preference outcomes:")
    for rank in sorted((r for r in counts if r is not None)):
        print(f"  rank {rank}: {counts[rank]}")
    print(f"  unlisted: {counts[None]}")
    print("Team sizes:")
    for p in sorted(teams):
        members = teams[p]
        n_msai = sum(1 for m in members if m["program"] == "MSAI")
        print(f"  {p:28} n={len(members)}  MSAI={n_msai} MBAi={len(members)-n_msai}  tech={TECH_RATING[p]}")


def gender_slice_note(teams) -> str:
    bits = []
    for project, members in sorted(teams.items()):
        for cohort in ("MSAI", "MBAi"):
            slice_ = [m for m in members if m["program"] == cohort]
            if len(slice_) != 3:
                continue
            n_f = sum(1 for m in slice_ if m.get("gender") == "f")
            bits.append(f"{project} {cohort} {len(slice_)-n_f}M/{n_f}F")
    return "3-person cohort slices: " + ("; ".join(bits) if bits else "none") + "."


def lone_female_note(teams) -> str:
    parts = []
    for project, members in sorted(teams.items()):
        n_f = sum(1 for m in members if m.get("gender") == "f")
        parts.append(f"{project} {n_f}F")
    zeros = [p for p, m in teams.items() if sum(1 for x in m if x.get("gender")=="f") == 0]
    return (
        "No team has exactly one female. Female counts: "
        + "; ".join(parts)
        + f". All-male teams: {', '.join(sorted(zeros)) or 'none'}."
    )


def even_gender_note(teams) -> str:
    parts = []
    for project, members in sorted(teams.items()):
        n_f = sum(1 for m in members if m.get("gender") == "f")
        n_m = len(members) - n_f
        parts.append(f"{project} {n_m}M/{n_f}F")
    counts = {}
    for members in teams.values():
        n_f = sum(1 for m in members if m.get("gender") == "f")
        counts[n_f] = counts.get(n_f, 0) + 1
    spread = ", ".join(f"{n}F×{c} teams" for n, c in sorted(counts.items()))
    n_f_total = sum(1 for members in teams.values() for m in members if m.get("gender") == "f")
    allowed = "0 or 2 women" if n_f_total % 2 == 0 else "0, 2, or 3 women"
    return (
        f"Gender is spread as evenly as {n_f_total} women across {len(teams)} teams allow without a lone female: "
        f"every team has {allowed} ({spread}). " + "; ".join(parts) + "."
    )


def write_trio(students, projects, sheets, titles, intros):
    variants = [
        dict(use_tech_mix=True),
        dict(use_tech_mix=True, forbid_2m1f=True),
        dict(use_tech_mix=True, forbid_lone_female=True),
    ]
    extras = [
        lambda teams: tech_mix_note(teams),
        lambda teams: tech_mix_note(teams) + " " + gender_slice_note(teams),
        lambda teams: tech_mix_note(teams) + " " + lone_female_note(teams),
    ]
    for (sheet, title, intro), kwargs, extra_fn in zip(zip(sheets, titles, intros), variants, extras):
        _, teams, objective = allocate(students, projects, **kwargs)
        write_sheet(
            WORKBOOK,
            sheet,
            title,
            students,
            projects,
            teams,
            objective,
            [intro] + describe_run(teams, extra=extra_fn(teams)),
            show_tech=True,
            show_gender=True,
        )
        print_run(sheet, teams, objective)


def main():
    survey_students, projects = load_students(WORKBOOK)
    missing_tech = [p for p in projects if p not in TECH_RATING]
    if missing_tech:
        raise SystemExit(f"Missing tech ratings: {missing_tech}")

    def describe_pop(students, leftover, forced=None):
        n4, n5, n6 = choose_team_sizes(len(students), len(projects))
        wild = ", ".join(
            f"{s['name']} ({s['program']}, {s['gender'].upper()})"
            for s in students if s.get("wildcard")
        )
        extra = ", ".join(
            f"{s['name']} ({s['program']}, {(s.get('gender') or '?').upper()})"
            for s in students if s.get("unrostered")
        )
        n_msai = sum(1 for s in students if s["program"] == "MSAI")
        n_mbai = sum(1 for s in students if s["program"] == "MBAi")
        print(f"Population {len(students)} (MSAI {n_msai}, MBAi {n_mbai}); "
              f"mix {n4} fours / {n5} fives / {n6} sixes")
        parts = []
        if n4:
            parts.append(f"{n4} team of 4" if n4 == 1 else f"{n4} teams of 4")
        if n5:
            parts.append(f"{n5} teams of 5")
        if n6:
            parts.append(f"{n6} teams of 6")
        size_txt = " and ".join(parts)
        forced = FORCED if forced is None else forced
        forced_txt = "; ".join(
            f"{person} is forced to {company} and shown in white"
            for person, company in forced.items()
        )
        intro = (
            "Same requirements as the tech-mix allocation, using the official class rosters. "
            f"Wildcards with no bid (cost 0): {wild}. "
            + (f"Included though not on a roster: {extra}. " if extra else
               f"Survey responses not on a roster were excluded: {', '.join(s['name'] for s in leftover) or 'none'}. ")
            + f"{len(students)} students → {size_txt}. {forced_txt}."
        )
        return intro

    with_amitha, _, leftover_a = merge_roster_and_survey(
        survey_students, projects, include_unrostered=True
    )

    erik_fee = {"Erik Beitel": "Future Energy Enterprises"}
    _, teams, obj = allocate(
        with_amitha, projects, use_tech_mix=True, forced=erik_fee
    )
    write_sheet(
        WORKBOOK,
        SHEET_V7,
        "Capstone allocations — gender-agnostic (+ Amitha; Erik → Future Energy)",
        with_amitha,
        projects,
        teams,
        obj,
        [
            describe_pop(with_amitha, leftover_a, forced=erik_fee)
            + " This tab does not use gender in the assignment. Amitha Javare Gowda is included as MSAI, gender F "
            "(inferred; she is not on the class roster).",
        ]
        + describe_run(teams, extra=tech_mix_note(teams), forced=erik_fee),
        show_tech=True,
        show_gender=True,
    )
    print_run(SHEET_V7, teams, obj)

    wb = load_workbook(WORKBOOK)
    order = [
        SHEET_IN,
        SHEET_V4, SHEET_V5, SHEET_V6, SHEET_V10, SHEET_V12,
        SHEET_V7, SHEET_V8, SHEET_V9, SHEET_V11, SHEET_V13,
    ]
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    wb.save(WORKBOOK)
    print("sheets:", load_workbook(WORKBOOK).sheetnames)


if __name__ == "__main__":
    main()
