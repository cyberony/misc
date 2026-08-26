#!/usr/bin/env python3
"""
Add "Classwork 1 Timely" column to the grade sheet (xlsx).
Class was 11:00 AM–12:30 PM; deadline = Jan 12, 2026 12:30 PM Central.
Uses GitHub API: repo created_at (when the repo was forked/created) to determine Timely vs Late.
Set USE_GITHUB_FORK_TIME=1 to use fork time; otherwise uses local submission_timestamps.txt (last commit).
Run from sandbox: python grading/add_classwork1_timely.py
"""
import csv
import json
import os
import time
import urllib.request
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

SANDBOX = Path(__file__).resolve().parent.parent
GRADING = SANDBOX / "grading"
LINKING_CSV = SANDBOX / "linking_table.csv"
GRADES_XLSX = GRADING / "2026-03-16T1228_Grades-2026WI_MSAI_371-0_SEC20.xlsx"
CW1_TIMESTAMPS = GRADING / "classwork-1-fopc" / "submission_timestamps.txt"

GITHUB_ORG = "NU-MSAI-371"
CW1_REPO_PREFIX = "classwork-1-fopc-"

# End of class: Jan 12, 2026 12:30 PM Central
DEADLINE_CENTRAL = datetime(2026, 1, 12, 12, 30, 0, tzinfo=ZoneInfo("America/Chicago"))

CW1_COL = "Classwork 1 - FOPC (1709355)"
TIMELY_COL = "Classwork 1 Timely"
MINUTES_LATE_COL = "Classwork 1 Minutes Late"
STUDENT_COL = "Student"
SIS_COL = "SIS User ID"


def load_linking():
    sis_to_gh = {}
    with open(LINKING_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            sis = (r.get("SIS User ID") or "").strip().upper()
            gh = (r.get("GitHub ID") or "").strip()
            if sis and gh:
                sis_to_gh[sis] = gh
    return sis_to_gh


def fetch_repo_created_at(ghid: str) -> datetime | None:
    """GET GitHub API repos/{org}/{repo}, return created_at in Central or None if 404/error."""
    repo = CW1_REPO_PREFIX + ghid
    url = f"https://api.github.com/repos/{GITHUB_ORG}/{repo}"
    req = urllib.request.Request(url)
    req.add_header("Accept", "application/vnd.github.v3+json")
    token = os.environ.get("GITHUB_TOKEN")
    if not token and (GRADING / ".github_token").exists():
        try:
            token = (GRADING / ".github_token").read_text().splitlines()[0].strip()
        except Exception:
            token = None
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise
    created = data.get("created_at")
    if not created:
        return None
    # GitHub returns "2026-01-12T18:30:00Z"
    dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
    return dt.astimezone(ZoneInfo("America/Chicago"))


def minutes_late(sub_central: datetime) -> int:
    """Minutes after deadline (0 if on time or early)."""
    delta = sub_central - DEADLINE_CENTRAL
    return max(0, int(round(delta.total_seconds() / 60)))


def get_timely_from_github(
    gh_ids: list[str],
    fallback_status: dict[str, str] | None = None,
    fallback_minutes: dict[str, int] | None = None,
) -> tuple[dict[str, str], dict[str, int]]:
    """Use GitHub API repo created_at (fork time). Returns (gh_to_status, gh_to_minutes)."""
    gh_to_status = {}
    gh_to_minutes = {}
    for i, ghid in enumerate(gh_ids):
        if i and i % 10 == 0:
            print(f"  Fetched {i}/{len(gh_ids)} repos...")
        if i > 0:
            time.sleep(0.6)  # avoid rate limit when unauthenticated
        try:
            created_central = fetch_repo_created_at(ghid)
        except Exception as e:
            if fallback_status and ghid in fallback_status:
                gh_to_status[ghid] = fallback_status[ghid]
                gh_to_minutes[ghid] = fallback_minutes.get(ghid, 0) if fallback_minutes else 0
            else:
                print(f"  Warning: {ghid}: {e}")
                gh_to_status[ghid] = ""
                gh_to_minutes[ghid] = 0
            continue
        if created_central is None:
            gh_to_status[ghid] = fallback_status.get(ghid, "") if fallback_status else ""
            gh_to_minutes[ghid] = fallback_minutes.get(ghid, 0) if fallback_minutes else 0
        else:
            gh_to_status[ghid] = "Timely" if created_central <= DEADLINE_CENTRAL else "Late"
            gh_to_minutes[ghid] = minutes_late(created_central)
    return gh_to_status, gh_to_minutes


def parse_classwork1_timestamps() -> tuple[dict[str, str], dict[str, int]]:
    """Return (gh_to_status, gh_to_minutes) from local submission_timestamps.txt (last commit time)."""
    gh_to_status = {}
    gh_to_minutes = {}
    with open(CW1_TIMESTAMPS, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("Submission") or line.startswith("Roster") or line.startswith("Deadline"):
                continue
            parts = line.split()
            if len(parts) < 2:
                continue
            repo_name = parts[0]
            if not repo_name.startswith("classwork-1-fopc-"):
                continue
            ghid = repo_name[len("classwork-1-fopc-"):]
            try:
                dt = datetime.fromisoformat(parts[1])
            except ValueError:
                continue
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=ZoneInfo("America/Chicago"))
            sub_central = dt.astimezone(ZoneInfo("America/Chicago"))
            gh_to_status[ghid] = "Timely" if sub_central <= DEADLINE_CENTRAL else "Late"
            gh_to_minutes[ghid] = minutes_late(sub_central)
    return gh_to_status, gh_to_minutes


def main():
    use_github = os.environ.get("USE_GITHUB_FORK_TIME", "1").strip().lower() in ("1", "true", "yes")

    xlsx_path = GRADES_XLSX
    if not xlsx_path.exists():
        for f in GRADING.glob("*.xlsx"):
            xlsx_path = f
            break
    if not xlsx_path.exists():
        print("No grade sheet .xlsx found in grading/.")
        return

    sis_to_gh = load_linking()
    if use_github:
        print("Using GitHub API: repo created_at (fork/creation time).")
        print("(Set GITHUB_TOKEN for higher rate limit. Fallback: last-commit time for failed fetches.)")
        gh_ids = list(sis_to_gh.values())
        fallback_status, fallback_minutes = parse_classwork1_timestamps()
        gh_to_timely, gh_to_minutes = get_timely_from_github(
            gh_ids, fallback_status=fallback_status, fallback_minutes=fallback_minutes
        )
    else:
        print("Using local submission_timestamps.txt (last commit time).")
        gh_to_timely, gh_to_minutes = parse_classwork1_timestamps()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    max_col = ws.max_column
    headers = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c

    cw1_col = headers.get(CW1_COL)
    sis_col = headers.get(SIS_COL)
    if not sis_col:
        print("Missing SIS User ID column.")
        return

    timely_col_idx = headers.get(TIMELY_COL)
    if timely_col_idx is None:
        timely_col_idx = (cw1_col + 1) if cw1_col else (max_col + 1)
        ws.cell(row=1, column=timely_col_idx, value=TIMELY_COL)

    minutes_col_idx = headers.get(MINUTES_LATE_COL)
    if minutes_col_idx is None:
        minutes_col_idx = timely_col_idx + 1
        ws.cell(row=1, column=minutes_col_idx, value=MINUTES_LATE_COL)

    for row in range(2, ws.max_row + 1):
        sis_val = ws.cell(row=row, column=sis_col).value
        sis = (sis_val or "").strip().upper() if sis_val is not None else ""
        if not sis:
            continue
        ghid = sis_to_gh.get(sis)
        if ghid is None:
            continue
        status = gh_to_timely.get(ghid, "")
        mins = gh_to_minutes.get(ghid, 0)
        ws.cell(row=row, column=timely_col_idx, value=status)
        if status:
            ws.cell(row=row, column=minutes_col_idx, value=mins)
        else:
            ws.cell(row=row, column=minutes_col_idx, value="")

    wb.save(xlsx_path)
    print("Added/updated column", TIMELY_COL, "in", xlsx_path)
    print("Deadline: Jan 12, 2026 12:30 PM Central (end of class).")
    print("Source:", "GitHub repo created_at (fork time)" if use_github else "local last-commit timestamps")


if __name__ == "__main__":
    main()
