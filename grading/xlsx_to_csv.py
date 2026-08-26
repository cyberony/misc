#!/usr/bin/env python3
"""
Port values from the grade sheet xlsx into the grade sheet CSV.
- Output CSV has the same columns as the xlsx (in xlsx order), inserting any missing columns.
- Copy all values from xlsx into CSV, matched by SIS User ID.
- Leave "Classwork 2 - Reasoning (1710189)" empty in the CSV.
Run from sandbox: python grading/xlsx_to_csv.py
"""
import csv
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

GRADING = Path(__file__).resolve().parent
GRADES_XLSX = GRADING / "2026-03-16T2229_Grades-2026WI_MSAI_371-0_SEC20.xlsx"
GRADES_CSV = GRADING / "2026-03-16T2229_Grades-2026WI_MSAI_371-0_SEC20.csv"
SKIP_COLUMN = "Classwork 2 - Reasoning (1710189)"


def main():
    xlsx_path = GRADES_XLSX
    if not xlsx_path.exists():
        for f in GRADING.glob("*.xlsx"):
            xlsx_path = f
            break
    if not xlsx_path.exists() or not GRADES_CSV.exists():
        print("Need both xlsx and CSV in grading/.")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    # Column order and indices from xlsx (row 1)
    xlsx_col_names = []
    xlsx_header_to_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        name = str(v).strip() if v is not None else ""
        xlsx_col_names.append(name)
        if name:
            xlsx_header_to_col[name] = c

    sis_col = xlsx_header_to_col.get("SIS User ID")
    if not sis_col:
        print("xlsx has no SIS User ID column.")
        return

    sis_to_xlsx_row = {}
    for row in range(2, ws.max_row + 1):
        sis = ws.cell(row=row, column=sis_col).value
        if sis is not None:
            sis_to_xlsx_row[str(sis).strip().upper()] = row

    # Read CSV (preserve row order for data rows)
    with open(GRADES_CSV, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        lines = list(reader)

    if len(lines) < 4:
        print("CSV has too few lines.")
        return

    csv_header = lines[0]
    sis_idx_csv = next((i for i, h in enumerate(csv_header) if (h or "").strip() == "SIS User ID"), None)
    if sis_idx_csv is None:
        print("CSV has no SIS User ID column.")
        return

    # Output: same columns as xlsx (so CSV gets all columns including new ones)
    output_lines = []
    # Row 1: xlsx header
    output_lines.append(xlsx_col_names)
    # Rows 2 and 3: from xlsx
    row2 = []
    row3 = []
    for c in range(1, ws.max_column + 1):
        v2 = ws.cell(row=2, column=c).value
        v3 = ws.cell(row=3, column=c).value
        row2.append("" if v2 is None else str(v2))
        row3.append("" if v3 is None else str(v3))
    output_lines.append(row2)
    output_lines.append(row3)

    # Data rows: for each CSV data row, find xlsx row by SIS User ID; output one row with all xlsx columns (values from xlsx, SKIP_COLUMN empty)
    for i in range(3, len(lines)):
        csv_row = lines[i]
        sis_val = csv_row[sis_idx_csv] if sis_idx_csv < len(csv_row) else ""
        sis = (sis_val or "").strip().upper()
        xlsx_row_num = sis_to_xlsx_row.get(sis)

        out_row = []
        for col_name in xlsx_col_names:
            if col_name == SKIP_COLUMN:
                out_row.append("")
                continue
            col_idx = xlsx_header_to_col.get(col_name)
            if col_idx is None or xlsx_row_num is None:
                out_row.append("")
                continue
            val = ws.cell(row=xlsx_row_num, column=col_idx).value
            out_row.append("" if val is None else str(val))
        output_lines.append(out_row)

    with open(GRADES_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for line in output_lines:
            writer.writerow(line)

    print("Port complete:", GRADES_CSV)
    print("Columns match xlsx (new columns inserted). Classwork 2 - Reasoning (1710189) left empty.")


if __name__ == "__main__":
    main()
