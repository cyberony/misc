#!/usr/bin/env python3
"""
Set Classwork 1 and Classwork 2 scores independently: 3 points if Timely for that classwork.
Run from sandbox: python grading/set_timely_classwork_scores.py
"""
from pathlib import Path

import openpyxl

GRADING = Path(__file__).resolve().parent
XLSX = GRADING / "2026-03-16T2229_Grades-2026WI_MSAI_371-0_SEC20.xlsx"

CW1_SCORE = "Classwork 1 - FOPC (1709355)"
CW2_SCORE = "Classwork 2 - Reasoning (1710189)"
CW1_TIMELY = "Classwork 1 Timely"
CW2_TIMELY = "Classwork 2 Timely"
STUDENT = "Student"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c

    col = lambda name: headers.get(name)
    data_start = 4
    c1_updated = []
    c2_updated = []

    for r in range(data_start, ws.max_row + 1):
        student = ws.cell(row=r, column=col(STUDENT)).value
        if student and "Test" in str(student):
            continue
        c1_timely = str(ws.cell(row=r, column=col(CW1_TIMELY)).value or "").strip() == "Timely"
        c2_timely = str(ws.cell(row=r, column=col(CW2_TIMELY)).value or "").strip() == "Timely"
        if c1_timely:
            ws.cell(row=r, column=col(CW1_SCORE), value=3)
            c1_updated.append((r, student))
        if c2_timely:
            ws.cell(row=r, column=col(CW2_SCORE), value=3)
            c2_updated.append((r, student))

    wb.save(XLSX)
    print(f"Classwork 1: set to 3 for {len(c1_updated)} timely students.")
    print(f"Classwork 2: set to 3 for {len(c2_updated)} timely students.")


if __name__ == "__main__":
    main()
