#!/usr/bin/env python3
"""
Apply Homework 1 late penalty to the grade sheet (xlsx) and ensure penalty notes in comments.
Expects: grading/2026-03-16T2229_Grades-2026WI_MSAI_371-0_SEC20.xlsx (or 1228)
Uses: linking_table.csv, homework-1-fopc-and-tables/submission_timestamps.txt
Penalty: 1 day = 10%, 2 = 25%, 3 = 50%, 4+ = 100%. Score becomes score * (1 - penalty/100).
Run from sandbox: pip install openpyxl && python grading/apply_hw1_late_penalty.py
"""
import csv
import re
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

SANDBOX = Path(__file__).resolve().parent.parent
GRADING = SANDBOX / "grading"
LINKING_CSV = SANDBOX / "linking_table.csv"
GRADES_XLSX = GRADING / "2026-03-16T2229_Grades-2026WI_MSAI_371-0_SEC20.xlsx"
TIMESTAMPS_FILE = GRADING / "homework-1-fopc-and-tables" / "submission_timestamps.txt"

HW1_COL = "Homework 1 - FOPC and Tables (1710992)"
COMMENTS_COL = "Homework 1 Comments"
DAYS_LATE_COL = "Homework 1 Days Late"
PENALTY_COL = "Homework 1 Penalty"
STUDENT_COL = "Student"
SIS_COL = "SIS User ID"

# End of day Jan 25, 2026 Central (Chicago)
DEADLINE_CENTRAL = datetime(2026, 1, 25, 23, 59, 59, tzinfo=ZoneInfo("America/Chicago"))
DEADLINE_UTC = DEADLINE_CENTRAL.astimezone(timezone.utc)
PENALTY_PCT = {1: 10, 2: 25, 3: 50}  # 4+ -> 100

# Late penalty exempt (e.g. approved extension). Map GitHub ID -> HW1 raw score to set.
HW1_EXEMPT = {"shi1gesong": 8.0}  # Li, Xiang


def load_linking():
    sis_to_gh = {}
    with open(LINKING_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            sis = (r.get("SIS User ID") or "").strip().upper()
            gh = (r.get("GitHub ID") or "").strip()
            if sis and gh:
                sis_to_gh[sis] = gh
    return sis_to_gh


def parse_timestamps():
    gh_to_days = {}
    gh_to_penalty = {}
    with open(TIMESTAMPS_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("Submission") or line.startswith("Roster") or line.startswith("Deadline"):
                continue
            parts = line.split()
            if len(parts) < 2:
                continue
            repo_name = parts[0]
            if not repo_name.startswith("homework-1-fopc-and-tables-"):
                continue
            ghid = repo_name[len("homework-1-fopc-and-tables-"):]
            if ghid in HW1_EXEMPT:
                gh_to_days[ghid] = 0
                gh_to_penalty[ghid] = 0
                continue
            try:
                dt = datetime.fromisoformat(parts[1])
            except ValueError:
                continue
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            sub_utc = dt.astimezone(timezone.utc)
            sub_central = sub_utc.astimezone(ZoneInfo("America/Chicago"))
            if sub_central <= DEADLINE_CENTRAL:
                days_late = 0
                penalty_pct = 0
            else:
                days_late = (sub_central.date() - DEADLINE_CENTRAL.date()).days
                penalty_pct = PENALTY_PCT.get(days_late, 100)
            gh_to_days[ghid] = days_late
            gh_to_penalty[ghid] = penalty_pct
    return gh_to_days, gh_to_penalty


def main():
    xlsx_path = GRADES_XLSX
    if not xlsx_path.exists():
        # Use any .xlsx in grading folder
        for f in GRADING.glob("*.xlsx"):
            xlsx_path = f
            break
    if not xlsx_path.exists():
        print("Expected grade sheet at:", GRADES_XLSX)
        print("Or place any .xlsx grade sheet in the grading/ folder and run again.")
        return

    sis_to_gh = load_linking()
    gh_to_days, gh_to_penalty = parse_timestamps()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    max_col = ws.max_column
    headers = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c

    hw1_col = headers.get(HW1_COL)
    comments_col = headers.get(COMMENTS_COL)
    days_col = headers.get(DAYS_LATE_COL)
    penalty_col = headers.get(PENALTY_COL)
    student_col = headers.get(STUDENT_COL)
    sis_col = headers.get(SIS_COL)

    if not all([hw1_col, comments_col, student_col, sis_col]):
        print("Missing columns. Found:", list(headers.keys())[:15], "...")
        return

    # If Days Late / Penalty columns missing, add them at end
    if days_col is None or penalty_col is None:
        days_col = max_col + 1
        penalty_col = max_col + 2
        ws.cell(row=1, column=days_col, value=DAYS_LATE_COL)
        ws.cell(row=1, column=penalty_col, value=PENALTY_COL)

    penalty_note_by_sis = {}
    for row in range(2, ws.max_row + 1):
        student = ws.cell(row=row, column=student_col).value
        if student is None:
            continue
        student = str(student).strip().strip('"')
        if student == "Student, Test":
            continue
        sis_val = ws.cell(row=row, column=sis_col).value
        sis = (sis_val or "").strip().upper() if sis_val is not None else ""
        if not sis:
            continue
        ghid = sis_to_gh.get(sis)
        if ghid is None:
            continue
        days_late = gh_to_days.get(ghid, 0)
        penalty_pct = gh_to_penalty.get(ghid, 0)
        penalty_str = f"{penalty_pct}%" if penalty_pct else "0%"

        # Exempt students (e.g. approved extension): set score and remove late note
        if ghid in HW1_EXEMPT:
            ws.cell(row=row, column=hw1_col, value=HW1_EXEMPT[ghid])
            comments_cell = ws.cell(row=row, column=comments_col)
            existing = (comments_cell.value or "").strip()
            existing = re.sub(
                r"\s*Late submission:\s*\d+\s+day(?:s)?(?:\(s\))?\s*\(\d+%\s*penalty applied\)\.?\s*",
                " ",
                existing,
                flags=re.IGNORECASE,
            ).strip()
            comments_cell.value = existing if existing else None
            ws.cell(row=row, column=days_col, value=0)
            ws.cell(row=row, column=penalty_col, value="0%")
            continue

        # If this student was previously marked late but is now on time, restore score and remove late note
        if days_late == 0 and penalty_col is not None:
            old_penalty_val = ws.cell(row=row, column=penalty_col).value
            old_penalty_pct = 0
            if old_penalty_val is not None:
                s = str(old_penalty_val).strip().replace("%", "")
                try:
                    old_penalty_pct = int(s)
                except ValueError:
                    pass
            if old_penalty_pct > 0:
                current_score_val = ws.cell(row=row, column=hw1_col).value
                try:
                    current_score = float(current_score_val)
                    raw_score = round(current_score / (1 - old_penalty_pct / 100), 2)
                    ws.cell(row=row, column=hw1_col, value=raw_score)
                except (TypeError, ValueError):
                    pass
            # Always remove late-submission line from comments when days_late is 0
            comments_cell = ws.cell(row=row, column=comments_col)
            existing = (comments_cell.value or "").strip()
            existing = re.sub(
                r"\s*Late submission:\s*\d+\s+day(?:s)?(?:\(s\))?\s*\(\d+%\s*penalty applied\)\.?\s*",
                " ",
                existing,
                flags=re.IGNORECASE,
            ).strip()
            comments_cell.value = existing if existing else None

        ws.cell(row=row, column=days_col, value=days_late)
        ws.cell(row=row, column=penalty_col, value=penalty_str)

        if days_late > 0:
            note = f"Late submission: {days_late} day(s) ({penalty_str} penalty applied)."
            penalty_note_by_sis[sis] = (days_late, penalty_str)
            comments_cell = ws.cell(row=row, column=comments_col)
            existing = (comments_cell.value or "").strip()
            if note not in existing:
                comments_cell.value = (existing + " " + note).strip() if existing else note

        if days_late > 0 and penalty_pct > 0:
            raw = ws.cell(row=row, column=hw1_col).value
            try:
                raw_score = float(raw)
            except (TypeError, ValueError):
                continue
            new_score = round(raw_score * (1 - penalty_pct / 100), 2)
            ws.cell(row=row, column=hw1_col, value=new_score)

    wb.save(xlsx_path)
    print("Updated", xlsx_path, "with late penalties applied to Homework 1 scores and comments.")


if __name__ == "__main__":
    main()
