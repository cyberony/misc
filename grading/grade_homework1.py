#!/usr/bin/env python3
"""
Grade Homework 1 (FOPC + Tables) for all 45 students.
Run from sandbox dir: python grading/grade_homework1.py
Updates the grade sheet: uses .xlsx if present, else .csv.
Ignores "Student, Test" (46th row).
"""
import csv
import subprocess
import sys
from pathlib import Path

SANDBOX = Path(__file__).resolve().parent.parent
GRADING = SANDBOX / "grading"
HOMEWORK_DIR = GRADING / "homework-1-fopc-and-tables"
LINKING_CSV = SANDBOX / "linking_table.csv"
GRADES_BASE = "2026-03-16T1228_Grades-2026WI_MSAI_371-0_SEC20"
GRADES_XLSX = GRADING / f"{GRADES_BASE}.xlsx"
GRADES_CSV = GRADING / f"{GRADES_BASE}.csv"
HW1_COL = "Homework 1 - FOPC and Tables (1710992)"
COMMENTS_COL = "Homework 1 Comments"
TOTAL_POINTS = 15  # FOPC 7 + Tables 8
TIMEOUT = 60


def load_linking():
    """SIS User ID (upper) -> GitHub ID."""
    sis_to_gh = {}
    with open(LINKING_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            sis = (r.get("SIS User ID") or "").strip().upper()
            gh = (r.get("GitHub ID") or "").strip()
            if sis and gh:
                sis_to_gh[sis] = gh
    return sis_to_gh


def grade_fopc(repo_path: Path) -> tuple[float, list[str]]:
    """Run homework_fopc.py and check key results. Returns (points 0-7, list of deduction reasons)."""
    points = 7.0
    comments = []
    py_file = repo_path / "homework_fopc.py"
    csv_file = repo_path / "honeyproduction.csv"
    if not py_file.exists():
        return 0.0, ["homework_fopc.py missing"]
    if not csv_file.exists():
        return 0.0, ["honeyproduction.csv missing"]

    # Run and capture
    try:
        r = subprocess.run(
            [sys.executable, "homework_fopc.py"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=TIMEOUT,
        )
    except subprocess.TimeoutExpired:
        return 0.0, ["FOPC timed out"]
    except Exception as e:
        return 0.0, [f"FOPC run error: {e}"]

    if r.returncode != 0:
        err = (r.stderr or r.stdout or "")[:300]
        return 0.0, [f"FOPC failed (exit {r.returncode}): {err}"]

    out = (r.stdout or "") + (r.stderr or "")
    # Check expected content
    if "626" not in out and "Loaded" not in out:
        points -= 1.0
        comments.append("load/encode: expected 626 records or Loaded summary")
    if "Major producer" not in out and "MajorProducer" not in out:
        points -= 1.0
        comments.append("major producers query missing or wrong")
    if "High price" not in out and "HighPrice" not in out:
        points -= 1.0
        comments.append("high price query missing or wrong")
    if "Production for" not in out and "Produced" not in out:
        points -= 1.0
        comments.append("production query missing or wrong")
    if "valid production" not in out.lower() and "ValidProduction" not in out:
        points -= 1.0
        comments.append("production constraint / ValidProduction missing or wrong")

    # If we already deducted 5, cap at 2 more for minor issues; else check for 27470000 (CA 2010)
    if points > 2 and "27470000" not in out and "27,470,000" not in out:
        points -= 0.5
        comments.append("CA production value not found or wrong")
    return max(0, points), comments


def grade_tables(repo_path: Path) -> tuple[float, list[str]]:
    """Run homework_tables.py. Returns (points 0-8, list of deduction reasons)."""
    points = 8.0
    comments = []
    py_file = repo_path / "homework_tables.py"
    csv_file = repo_path / "honeyproduction.csv"
    if not py_file.exists():
        return 0.0, ["homework_tables.py missing"]
    if not csv_file.exists():
        return 0.0, ["honeyproduction.csv missing"]

    try:
        r = subprocess.run(
            [sys.executable, "homework_tables.py"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=TIMEOUT,
        )
    except subprocess.TimeoutExpired:
        return 0.0, ["Tables timed out"]
    except Exception as e:
        return 0.0, [f"Tables run error: {e}"]

    if r.returncode != 0:
        err = (r.stderr or r.stdout or "")[:300]
        return 0.0, [f"Tables failed (exit {r.returncode}): {err}"]

    out = (r.stdout or "") + (r.stderr or "")
    # Minimal check: ran and produced some output (exercises 2-5)
    if "exercise" not in out.lower() and "Exercise" not in out and len(out.strip()) < 200:
        points -= 2.0
        comments.append("Tables: little or no exercise output")
    return max(0, points), comments


def main():
    sis_to_gh = load_linking()
    results = {}  # (Student, SIS User ID) -> (score, comments_str)
    for _sis, ghid in sis_to_gh.items():
        student = None  # resolve from sheet
        repo_path = HOMEWORK_DIR / f"homework-1-fopc-and-tables-{ghid}"
        if not repo_path.is_dir():
            continue
        fopc_pts, fopc_c = grade_fopc(repo_path)
        tab_pts, tab_c = grade_tables(repo_path)
        total = fopc_pts + tab_pts
        all_comments = []
        if fopc_c:
            all_comments.append("FOPC: " + "; ".join(fopc_c))
        if tab_c:
            all_comments.append("Tables: " + "; ".join(tab_c))
        if total < TOTAL_POINTS and not all_comments:
            all_comments.append(f"FOPC {fopc_pts}/7, Tables {tab_pts}/8")
        results[(_sis, ghid)] = (total, " | ".join(all_comments) if all_comments else "")

    use_xlsx = GRADES_XLSX.exists()
    if use_xlsx:
        try:
            import openpyxl
        except ImportError:
            print("xlsx found but openpyxl missing. Install: pip install openpyxl")
            use_xlsx = False
    if use_xlsx:
        wb = openpyxl.load_workbook(GRADES_XLSX)
        ws = wb.active
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None:
                headers[str(v).strip()] = c
        hw1_col = headers.get(HW1_COL)
        comments_col = headers.get(COMMENTS_COL)
        sis_col = headers.get("SIS User ID")
        if not all([hw1_col, comments_col, sis_col]):
            print("Missing columns in xlsx. Falling back to CSV if present.")
            use_xlsx = False
        else:
            for row in range(2, ws.max_row + 1):
                sis_val = ws.cell(row=row, column=sis_col).value
                sis = (sis_val or "").strip().upper() if sis_val is not None else ""
                if not sis or sis == "STUDENT, TEST":
                    continue
                ghid = sis_to_gh.get(sis)
                if ghid is None:
                    continue
                key = (sis, ghid)
                if key not in results:
                    continue
                score, comment = results[key]
                if score is not None:
                    ws.cell(row=row, column=hw1_col, value=round(score, 2))
                ws.cell(row=row, column=comments_col, value=comment)
            wb.save(GRADES_XLSX)
            print("Grading complete. Updated", GRADES_XLSX)
            print("Scores and comments applied for", len(results), "students (excluding Test Student).")
            return

    # CSV path
    if not GRADES_CSV.exists():
        print("No grade sheet found. Expected", GRADES_XLSX, "or", GRADES_CSV)
        return
    with open(GRADES_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)
    if COMMENTS_COL not in fieldnames:
        idx = fieldnames.index(HW1_COL) + 1
        fieldnames.insert(idx, COMMENTS_COL)
    for row in rows:
        student = (row.get("Student") or "").strip().strip('"')
        sis = (row.get("SIS User ID") or "").strip().upper()
        if student == "Student, Test" or not sis:
            continue
        ghid = sis_to_gh.get(sis)
        if ghid is None:
            continue
        key = (sis, ghid)
        if key not in results:
            continue
        score, comment = results[key]
        if score is not None:
            row[HW1_COL] = f"{score:.2f}"
        row[COMMENTS_COL] = comment
    with open(GRADES_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow(row)
    print("Grading complete. Updated", GRADES_CSV)
    print("Scores and comments applied for", len(results), "students (excluding Test Student).")


if __name__ == "__main__":
    main()
